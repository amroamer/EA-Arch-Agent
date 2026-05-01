"""CRUD for EA compliance frameworks (Settings page).

Frameworks are simple aggregates: a Framework owns a list of items.
The PUT endpoint replaces the entire items list atomically so the
frontend can save the whole table in one request without tracking
per-row dirty state.

`GET /frameworks/export.xlsx` produces a single workbook with one
sheet per framework, matching the layout of the source spreadsheets the
user originally OCR'd in (Cover page + per-framework tabs).
"""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import desc, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.db import Framework, FrameworkItem
from app.models.framework_schemas import (
    FrameworkDetail,
    FrameworkItemBase,
    FrameworkSummary,
    FrameworkUpsert,
)

logger = logging.getLogger(__name__)
router = APIRouter()


# ── Helpers ────────────────────────────────────────────────────────────


def _replace_items(
    framework: Framework, items: list[FrameworkItemBase]
) -> None:
    """Replace the framework's items list. SQLAlchemy will issue
    DELETE + INSERTs against the cascade='delete-orphan' relationship."""
    framework.items = [
        FrameworkItem(
            criteria=it.criteria,
            weight_planned=it.weight_planned,
            weight_actual=it.weight_actual,
            compliance_pct=it.compliance_pct,
            remarks=it.remarks,
            sort_order=idx if it.sort_order == 0 else it.sort_order,
        )
        for idx, it in enumerate(items)
    ]


# ── Routes ─────────────────────────────────────────────────────────────


@router.get("/frameworks", response_model=list[FrameworkSummary])
async def list_frameworks(
    db: AsyncSession = Depends(get_db),
) -> list[FrameworkSummary]:
    """All frameworks, newest first, with item counts."""
    item_count = (
        select(
            FrameworkItem.framework_id,
            func.count(FrameworkItem.id).label("count"),
        )
        .group_by(FrameworkItem.framework_id)
        .subquery()
    )
    stmt = (
        select(Framework, func.coalesce(item_count.c.count, 0))
        .outerjoin(item_count, Framework.id == item_count.c.framework_id)
        .order_by(desc(Framework.updated_at))
    )
    rows = (await db.execute(stmt)).all()
    return [
        FrameworkSummary(
            id=fw.id,
            name=fw.name,
            description=fw.description,
            created_at=fw.created_at,
            updated_at=fw.updated_at,
            item_count=int(count),
        )
        for fw, count in rows
    ]


@router.post("/frameworks", response_model=FrameworkDetail, status_code=201)
async def create_framework(
    body: FrameworkUpsert,
    db: AsyncSession = Depends(get_db),
) -> FrameworkDetail:
    fw = Framework(name=body.name.strip(), description=body.description)
    _replace_items(fw, body.items)
    db.add(fw)
    await db.commit()
    await db.refresh(fw, attribute_names=["items"])
    return FrameworkDetail.model_validate(fw)


# ── Excel export ──────────────────────────────────────────────────────
# IMPORTANT: must be registered BEFORE /frameworks/{framework_id} so the
# literal path doesn't get matched against the path-parameter route.


_INVALID_SHEET_CHARS = re.compile(r"[\\/?*\[\]:]")


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: max 31 chars, no `: \\ / ? * [ ]`, must be unique."""
    safe = _INVALID_SHEET_CHARS.sub("-", name).strip() or "Framework"
    if len(safe) > 31:
        safe = safe[:31]
    base = safe
    counter = 2
    while safe.lower() in used:
        suffix = f" ({counter})"
        safe = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(safe.lower())
    return safe


def _build_workbook(frameworks: list[Framework]) -> bytes:
    """Render frameworks into a multi-sheet xlsx and return the bytes.

    Layout per sheet:
        A1     : Framework title (merged across all columns, bold, large)
        A3..F3 : Column headers
        A4+    : Criteria rows
        last   : Totals row (Weight columns summed)

    A leading "Cover Page" sheet summarizes the workbook.
    """
    wb = Workbook()
    # Drop the auto-created default sheet.
    wb.remove(wb.active)

    # Shared styles.
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="0C233C")  # kpmg darkBlue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="00338D")  # kpmg blue
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    num_align = Alignment(horizontal="right", vertical="top")
    totals_font = Font(name="Calibri", size=11, bold=True)
    totals_fill = PatternFill("solid", fgColor="ACEAFF")  # kpmg lightBlue
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Cover sheet ──
    cover = wb.create_sheet("Cover Page")
    cover["A1"] = "EA Compliance Frameworks"
    cover["A1"].font = Font(name="Calibri", size=18, bold=True, color="0C233C")
    cover["A2"] = (
        f"Exported {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        f"  ·  {len(frameworks)} framework(s)"
    )
    cover["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    cover.append([])  # blank row 3
    cover_headers = ["#", "Framework", "Criteria count", "Total planned weight"]
    for col, h in enumerate(cover_headers, start=1):
        c = cover.cell(row=4, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border
    for i, fw in enumerate(frameworks, start=1):
        total_planned = sum(float(it.weight_planned or 0) for it in fw.items)
        row = 4 + i
        cover.cell(row=row, column=1, value=i).alignment = num_align
        cover.cell(row=row, column=2, value=fw.name).alignment = body_align
        cover.cell(row=row, column=3, value=len(fw.items)).alignment = num_align
        cover.cell(row=row, column=4, value=round(total_planned, 2)).alignment = num_align
        for col in range(1, 5):
            cover.cell(row=row, column=col).border = border
    cover.column_dimensions["A"].width = 5
    cover.column_dimensions["B"].width = 45
    cover.column_dimensions["C"].width = 16
    cover.column_dimensions["D"].width = 22

    # ── Per-framework sheets ──
    used: set[str] = {"cover page"}
    columns = [
        ("#", 5),
        ("Criteria", 70),
        ("Weight (planned)", 16),
        ("Weight (actual)", 16),
        ("Compliance %", 14),
        ("Remarks", 40),
    ]

    for fw in frameworks:
        ws = wb.create_sheet(_sanitize_sheet_name(fw.name, used))

        # Title row spans all columns.
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=fw.name)
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        # Header row.
        for col_idx, (h, w) in enumerate(columns, start=1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.row_dimensions[3].height = 32

        # Criteria rows.
        items = sorted(fw.items, key=lambda it: it.sort_order)
        total_planned = 0.0
        total_actual = 0.0
        for i, it in enumerate(items, start=1):
            r = 3 + i
            wp = float(it.weight_planned or 0)
            wa = float(it.weight_actual or 0)
            total_planned += wp
            total_actual += wa
            ws.cell(row=r, column=1, value=i).alignment = num_align
            ws.cell(row=r, column=2, value=it.criteria).alignment = body_align
            ws.cell(row=r, column=3, value=wp).alignment = num_align
            ws.cell(row=r, column=4, value=wa).alignment = num_align
            ws.cell(row=r, column=5, value=float(it.compliance_pct or 0)).alignment = num_align
            ws.cell(row=r, column=6, value=it.remarks or "").alignment = body_align
            for col in range(1, len(columns) + 1):
                ws.cell(row=r, column=col).border = border

        # Totals row.
        if items:
            r = 3 + len(items) + 1
            label = ws.cell(row=r, column=2, value="Total")
            label.font = totals_font
            label.fill = totals_fill
            label.alignment = Alignment(horizontal="right", vertical="center")
            for col, val in [(3, round(total_planned, 2)), (4, round(total_actual, 2))]:
                c = ws.cell(row=r, column=col, value=val)
                c.font = totals_font
                c.fill = totals_fill
                c.alignment = num_align
            for col in range(1, len(columns) + 1):
                ws.cell(row=r, column=col).fill = totals_fill
                ws.cell(row=r, column=col).border = border

        # Freeze the title + header so scrolling keeps them visible.
        ws.freeze_panes = "A4"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@router.get("/frameworks/export.xlsx")
async def export_frameworks_xlsx(
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Download all frameworks as a single Excel workbook.

    One sheet per framework + a leading Cover Page summary sheet.
    Sheet names are derived from framework names (sanitized + uniquified).
    """
    stmt = (
        select(Framework)
        .order_by(Framework.name)
        .options(selectinload(Framework.items))
    )
    frameworks = list((await db.execute(stmt)).scalars().all())

    if not frameworks:
        raise HTTPException(404, "No frameworks to export")

    blob = _build_workbook(frameworks)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"EA-Compliance-Frameworks_{stamp}.xlsx"

    logger.info(
        "frameworks_exported",
        extra={"count": len(frameworks), "bytes": len(blob)},
    )

    return StreamingResponse(
        io.BytesIO(blob),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(blob)),
        },
    )


@router.get("/frameworks/{framework_id}", response_model=FrameworkDetail)
async def get_framework(
    framework_id: str,
    db: AsyncSession = Depends(get_db),
) -> FrameworkDetail:
    stmt = (
        select(Framework)
        .where(Framework.id == framework_id)
        .options(selectinload(Framework.items))
    )
    fw = (await db.execute(stmt)).scalar_one_or_none()
    if fw is None:
        raise HTTPException(404, "Framework not found")
    return FrameworkDetail.model_validate(fw)


@router.put("/frameworks/{framework_id}", response_model=FrameworkDetail)
async def update_framework(
    framework_id: str,
    body: FrameworkUpsert,
    db: AsyncSession = Depends(get_db),
) -> FrameworkDetail:
    """Full-update — replaces name, description, and the items list."""
    stmt = (
        select(Framework)
        .where(Framework.id == framework_id)
        .options(selectinload(Framework.items))
    )
    fw = (await db.execute(stmt)).scalar_one_or_none()
    if fw is None:
        raise HTTPException(404, "Framework not found")

    fw.name = body.name.strip()
    fw.description = body.description
    _replace_items(fw, body.items)
    await db.commit()
    await db.refresh(fw, attribute_names=["items"])
    return FrameworkDetail.model_validate(fw)


@router.delete("/frameworks/{framework_id}")
async def delete_framework(
    framework_id: str,
    db: AsyncSession = Depends(get_db),
) -> Response:
    fw = await db.get(Framework, framework_id)
    if fw is None:
        raise HTTPException(404, "Framework not found")
    await db.delete(fw)
    await db.commit()
    return Response(status_code=204)
